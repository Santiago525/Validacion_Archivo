# -*- coding: utf-8 -*-
"""
Created on Mon Aug  5 12:05:00 2024

@author: sagudelo
"""
# Importar libreria pandas 

import pandas as pd
from openpyxl import load_workbook


# leer archivo con los datos

#archivo = pd.read_excel('C:/Users/sagudelo/Documents/890939936_SOCIEDADMEDICARIONEGRO_20062024.xlsx',dtype=str)

# Leer archivo de la estructura
estructura = pd.read_excel('C:/Users/sagudelo/Documents/Estructura.xlsx')

estructura = dict(zip(estructura['CAMPO'], estructura['LONGITUD']))


#%%

# Función para leer el archivo Excel en chunks
def read_excel_in_chunks(file_path, chunk_size):
    workbook = load_workbook(filename=file_path, read_only=True)
    sheet = workbook.active

    header = [cell.value for cell in sheet[1]]
    chunk = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        chunk.append(row)
        if len(chunk) == chunk_size:
            yield pd.DataFrame(chunk, columns=header)
            chunk = []

    if chunk:
        yield pd.DataFrame(chunk, columns=header)

#Validar campos de acuerdo a la longitud en la estructura
#%%
errores = []

chunk_size = 100000  # Tamaño del chunk (puedes ajustarlo según tus necesidades)

archivo = 'C:/Users/sagudelo/Documents/890939936_Sociedad Medica Rionegro SA SOMER SA_2021.xlsx'

for chunk in read_excel_in_chunks(archivo,chunk_size):
   for columna in chunk.columns:
        if columna in estructura:
            longitud_max = estructura[columna]
            filas_malas = chunk[chunk[columna].astype(str).str.len()>longitud_max]
            for _,fila in filas_malas.iterrows():
                error = (
                    {
                      'FACTURA':fila.get('NUMERO DE FACTURA'),
                      'CODIGO SERVICIO': fila.get('CODIGO DEL SERVICIO FACTURADO'),
                      'DESCRIPCION DEL SERVICIO':fila.get('DECRIPCION DEL SERVICIO'),
                      'CAMPO A VALIDAR':columna,
                      'VALOR':fila[columna],
                      'LONGITUD':len(str(fila[columna])),
                      'LONGITUD PERMITIDA':longitud_max
                        }
                    )
                errores.append(error)

errores_df = pd.DataFrame(errores)

errores_df.to_excel('C:/Users/sagudelo/Documents/validacion.xlsx',index=False)

#%%

##Realizar correcion de cuando el campo cantidad es mayor de 5 digitos
archivo = pd.read_excel('C:/Users/sagudelo/Documents/890939936_Sociedad Medica Rionegro SA SOMER SA_2021.xlsx',dtype=str)


archivo['CANTIDAD']= archivo['CANTIDAD'].astype(int)
archivo['NUMERO DE FACTURA']= archivo['NUMERO DE FACTURA'].astype(int)
archivo['CONSECUTIVO DE LA FACTURA']= archivo['CONSECUTIVO DE LA FACTURA'].astype(int)

#%%
cantidad_mayores = archivo[archivo['CANTIDAD'] > 99999]


#%%

cantidades_menores = archivo[archivo['CANTIDAD']<=99999]


resultados =[]

for indice, fila in cantidad_mayores.iterrows():
    cantidad = fila['CANTIDAD']
    
    while cantidad > 99999:
        fila_nueva = fila.copy()
        fila_nueva['CANTIDAD'] = 99999
        resultados.append(fila_nueva)
        cantidad -= 99999
        
    if cantidad > 0:
            fila_nueva = fila.copy()
            fila_nueva['CANTIDAD'] = cantidad
            resultados.append(fila_nueva)
            
resultados = pd.DataFrame(resultados)

archivo_final = pd.concat([cantidades_menores,resultados],ignore_index=True)




archivo_final.to_excel('C:/Users/sagudelo/Documents/890939936_Sociedad Medica Rionegro SA SOMER SA_2023.xlsx', index=False)

#archivo_final.to_csv('C:/Users/sagudelo/Documents/Pruebadivision2.txt', sep='|', index=False)

print("Procesamiento y exportación a TXT completados.")



